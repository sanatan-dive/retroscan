"""
IRC compliance report generator.

Generates professional PDF and Excel reports for NHAI submission.
Supports English, Hindi, and bilingual rendering for Indian regulatory contexts.
"""
from __future__ import annotations

import io
import json
import os
import urllib.request
from datetime import datetime, timezone
from html import escape as html_escape
from pathlib import Path
from typing import Optional

from reportlab.lib import colors
from reportlab.lib.pagesizes import A4
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import mm, cm
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont
from reportlab.platypus import (
    SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, Image, PageBreak,
)
from reportlab.lib.enums import TA_CENTER, TA_LEFT

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

from app.services.database import get_scan, get_scan_detections, get_scan_frames
from app.config import RESULTS_DIR


# ---------------------------------------------------------------------------
# Hindi font registration
# ---------------------------------------------------------------------------

HINDI_FONT_NAME = "HindiSans"
HINDI_FONT_NAME_BOLD = "HindiSans-Bold"
_HINDI_FONT_REGISTERED = False
_HINDI_FONT_AVAILABLE = False

# Order of preference: macOS bundled fonts, then a downloaded Noto Sans Devanagari.
_HINDI_FONT_CANDIDATES = [
    "/System/Library/Fonts/Supplemental/NotoSansDevanagari.ttc",
    "/System/Library/Fonts/Supplemental/NotoSansDevanagari-Regular.ttf",
    "/System/Library/Fonts/Supplemental/Devanagari Sangam MN.ttc",
    "/System/Library/Fonts/Supplemental/DevanagariMT.ttc",
    "/System/Library/Fonts/Supplemental/ITFDevanagari.ttc",
    "/usr/share/fonts/truetype/noto/NotoSansDevanagari-Regular.ttf",
    "/usr/share/fonts/truetype/lohit-devanagari/Lohit-Devanagari.ttf",
]

_FONTS_DIR = Path(__file__).resolve().parent.parent.parent / "fonts"
_DOWNLOADED_FONT_PATH = _FONTS_DIR / "NotoSansDevanagari-Regular.ttf"
_DOWNLOADED_FONT_BOLD_PATH = _FONTS_DIR / "NotoSansDevanagari-Bold.ttf"

# Public CDN URLs (jsdelivr mirror of the Google Fonts repository)
_FONT_DOWNLOAD_URL = (
    "https://cdn.jsdelivr.net/gh/googlefonts/noto-fonts@main/"
    "hinted/ttf/NotoSansDevanagari/NotoSansDevanagari-Regular.ttf"
)
_FONT_DOWNLOAD_URL_BOLD = (
    "https://cdn.jsdelivr.net/gh/googlefonts/noto-fonts@main/"
    "hinted/ttf/NotoSansDevanagari/NotoSansDevanagari-Bold.ttf"
)


def _try_download_hindi_font() -> Path | None:
    """Best-effort download of Noto Sans Devanagari to backend/fonts/."""
    try:
        _FONTS_DIR.mkdir(parents=True, exist_ok=True)
        if not _DOWNLOADED_FONT_PATH.exists():
            urllib.request.urlretrieve(_FONT_DOWNLOAD_URL, _DOWNLOADED_FONT_PATH)
        if not _DOWNLOADED_FONT_BOLD_PATH.exists():
            try:
                urllib.request.urlretrieve(_FONT_DOWNLOAD_URL_BOLD, _DOWNLOADED_FONT_BOLD_PATH)
            except Exception:
                pass
        return _DOWNLOADED_FONT_PATH if _DOWNLOADED_FONT_PATH.exists() else None
    except Exception:
        return None


def _ensure_hindi_font() -> bool:
    """Register a Devanagari-capable TTF with ReportLab. Returns True on success."""
    global _HINDI_FONT_REGISTERED, _HINDI_FONT_AVAILABLE
    if _HINDI_FONT_REGISTERED:
        return _HINDI_FONT_AVAILABLE

    font_path: Path | None = None
    bold_path: Path | None = None

    for candidate in _HINDI_FONT_CANDIDATES:
        if os.path.exists(candidate):
            font_path = Path(candidate)
            break

    if font_path is None:
        downloaded = _try_download_hindi_font()
        if downloaded:
            font_path = downloaded
            if _DOWNLOADED_FONT_BOLD_PATH.exists():
                bold_path = _DOWNLOADED_FONT_BOLD_PATH

    if font_path is None:
        _HINDI_FONT_REGISTERED = True
        _HINDI_FONT_AVAILABLE = False
        return False

    try:
        # ReportLab's TTFont supports both .ttf and individual faces from .ttc
        # (subfontIndex defaults to 0 which is fine for our Devanagari fonts).
        if str(font_path).lower().endswith(".ttc"):
            pdfmetrics.registerFont(TTFont(HINDI_FONT_NAME, str(font_path), subfontIndex=0))
        else:
            pdfmetrics.registerFont(TTFont(HINDI_FONT_NAME, str(font_path)))

        if bold_path and bold_path.exists():
            pdfmetrics.registerFont(TTFont(HINDI_FONT_NAME_BOLD, str(bold_path)))
        else:
            # Fall back to using the regular face for bold runs.
            pdfmetrics.registerFont(TTFont(HINDI_FONT_NAME_BOLD, str(font_path),
                                           subfontIndex=0 if str(font_path).lower().endswith(".ttc") else 0))

        _HINDI_FONT_AVAILABLE = True
    except Exception:
        _HINDI_FONT_AVAILABLE = False

    _HINDI_FONT_REGISTERED = True
    return _HINDI_FONT_AVAILABLE


# ---------------------------------------------------------------------------
# Translation tables
# ---------------------------------------------------------------------------

T = {
    "report_title": {
        "en": "Retroreflectivity Compliance Report",
        "hi": "रेट्रोरिफ्लेक्टिविटी अनुपालन रिपोर्ट",
    },
    "survey_summary": {
        "en": "Survey Summary",
        "hi": "सर्वेक्षण सारांश",
    },
    "detailed_findings": {
        "en": "Detailed Asset Findings",
        "hi": "विस्तृत संपत्ति निष्कर्ष",
    },
    "recommendations": {
        "en": "Maintenance Recommendations",
        "hi": "रखरखाव की सिफारिशें",
    },
    "total_frames": {
        "en": "Total Frames Analyzed",
        "hi": "विश्लेषण किए गए कुल फ्रेम",
    },
    "total_assets": {
        "en": "Total Assets Detected",
        "hi": "कुल संपत्तियां पाई गईं",
    },
    "compliance_pass": {
        "en": "Compliance PASS",
        "hi": "अनुपालन उत्तीर्ण",
    },
    "compliance_fail": {
        "en": "Compliance FAIL",
        "hi": "अनुपालन विफल",
    },
    "signs_detected": {
        "en": "Signs Detected",
        "hi": "पाए गए संकेत",
    },
    "markings_detected": {
        "en": "Markings Detected",
        "hi": "पाए गए चिह्न",
    },
    "studs_detected": {
        "en": "Road Studs Detected",
        "hi": "पाए गए रोड स्टड",
    },
    "avg_score": {
        "en": "Average Retro Score",
        "hi": "औसत रेट्रो स्कोर",
    },
    "metric": {"en": "Metric", "hi": "मापदंड"},
    "value": {"en": "Value", "hi": "मूल्य"},
    "report_id": {"en": "Report ID", "hi": "रिपोर्ट आईडी"},
    "file_analyzed": {"en": "File Analyzed", "hi": "विश्लेषित फ़ाइल"},
    "date": {"en": "Date", "hi": "दिनांक"},
    "standards": {"en": "Standards", "hi": "मानक"},
    "conditions": {"en": "Conditions", "hi": "स्थितियां"},
    "asset_type": {"en": "Asset Type", "hi": "संपत्ति प्रकार"},
    "retro_class": {"en": "Retro Class", "hi": "रेट्रो श्रेणी"},
    "score": {"en": "Score", "hi": "स्कोर"},
    "grade": {"en": "Grade", "hi": "ग्रेड"},
    "irc_status": {"en": "IRC Status", "hi": "आईआरसी स्थिति"},
    "priority": {"en": "Priority", "hi": "प्राथमिकता"},
    "no_assets": {
        "en": "No assets detected in the analyzed footage.",
        "hi": "विश्लेषित फुटेज में कोई संपत्ति नहीं मिली।",
    },
    "no_critical": {
        "en": "No critical maintenance items found.",
        "hi": "कोई गंभीर रखरखाव आइटम नहीं मिला।",
    },
    "frame": {"en": "Frame", "hi": "फ्रेम"},
    "issues": {"en": "Issues", "hi": "समस्याएं"},
    "footer": {
        "en": "Report generated by RetroScan AI — AI-powered retroreflectivity assessment system. "
              "Standards referenced: IRC 67 (Traffic Signs), IRC 35 (Road Markings), IRC SP:79 (Road Studs).",
        "hi": "रिपोर्ट RetroScan AI द्वारा तैयार — एआई-संचालित रेट्रोरिफ्लेक्टिविटी मूल्यांकन प्रणाली। "
              "संदर्भित मानक: आईआरसी 67 (यातायात संकेत), आईआरसी 35 (सड़क चिह्न), आईआरसी एसपी:79 (रोड स्टड)।",
    },
}


def _label(key: str, language: str) -> str:
    """Return label in the requested language ('en', 'hi', or 'bilingual')."""
    en = T[key]["en"]
    hi = T[key]["hi"]
    if language == "hi":
        return hi
    if language == "bilingual":
        return f"{en} / {hi}"
    return en


def _para_text(text: str, language: str, hindi_available: bool) -> str:
    """Wrap text with the right font tag for ReportLab paragraphs."""
    safe = html_escape(str(text))
    if language == "en" or not hindi_available:
        return safe
    # Use registered Devanagari font for hi / bilingual content.
    return f'<font name="{HINDI_FONT_NAME}">{safe}</font>'


# ---------------------------------------------------------------------------
# PDF report
# ---------------------------------------------------------------------------

def generate_pdf_report(scan_id: str, language: str = "en") -> bytes:
    """Generate a PDF compliance report.

    Args:
        scan_id: scan identifier.
        language: 'en' (default), 'hi', or 'bilingual'.
    """
    language = (language or "en").lower()
    if language not in {"en", "hi", "bilingual"}:
        language = "en"

    scan = get_scan(scan_id)
    detections = get_scan_detections(scan_id)

    if not scan:
        raise ValueError(f"Scan {scan_id} not found")

    hindi_available = _ensure_hindi_font() if language in {"hi", "bilingual"} else False
    base_font = HINDI_FONT_NAME if (language == "hi" and hindi_available) else "Helvetica"
    base_font_bold = (
        HINDI_FONT_NAME_BOLD if (language == "hi" and hindi_available) else "Helvetica-Bold"
    )

    buffer = io.BytesIO()
    doc = SimpleDocTemplate(
        buffer, pagesize=A4,
        rightMargin=2 * cm, leftMargin=2 * cm,
        topMargin=2 * cm, bottomMargin=2 * cm,
    )

    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        "CustomTitle", parent=styles["Title"],
        fontSize=20, spaceAfter=10, textColor=colors.HexColor("#1a1a2e"),
        fontName=base_font_bold,
    )
    heading_style = ParagraphStyle(
        "CustomHeading", parent=styles["Heading2"],
        fontSize=14, spaceAfter=8, textColor=colors.HexColor("#16213e"),
        fontName=base_font_bold,
    )
    body_style = ParagraphStyle(
        "CustomBody", parent=styles["Normal"], fontName=base_font, fontSize=10,
    )

    elements = []

    # Title
    elements.append(Paragraph("RetroScan AI", title_style))
    elements.append(Paragraph(
        _para_text(_label("report_title", language), language, hindi_available),
        heading_style,
    ))
    elements.append(Spacer(1, 5 * mm))

    # Report metadata (handle both dict from JSONB and string)
    weather_raw = scan.get("weather_json", {})
    if isinstance(weather_raw, str):
        weather = json.loads(weather_raw or "{}")
    else:
        weather = weather_raw or {}
    cond_value = (
        f"{weather.get('time_of_day', 'N/A')} | "
        f"{weather.get('visibility', 'N/A')} | "
        f"{weather.get('moisture', 'N/A')}"
    )
    meta_pairs = [
        ("report_id", scan_id[:12]),
        ("file_analyzed", scan["filename"]),
        ("date", datetime.now(timezone.utc).strftime("%Y-%m-%d %H:%M UTC")),
        ("standards", "IRC 67 (Signs) / IRC 35 (Markings)"),
        ("conditions", cond_value),
    ]
    meta_data = [
        [
            Paragraph(_para_text(_label(key, language), language, hindi_available), body_style),
            Paragraph(_para_text(val, language, hindi_available), body_style),
        ]
        for key, val in meta_pairs
    ]
    meta_table = Table(meta_data, colWidths=[5 * cm, 11 * cm])
    meta_table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (0, -1), colors.HexColor("#e8eaf6")),
        ("FONTNAME", (0, 0), (0, -1), base_font_bold),
        ("FONTNAME", (1, 0), (1, -1), base_font),
        ("FONTSIZE", (0, 0), (-1, -1), 10),
        ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("LEFTPADDING", (0, 0), (-1, -1), 8),
        ("RIGHTPADDING", (0, 0), (-1, -1), 8),
        ("TOPPADDING", (0, 0), (-1, -1), 5),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 5),
    ]))
    elements.append(meta_table)
    elements.append(Spacer(1, 8 * mm))

    # Summary section
    elements.append(Paragraph(
        _para_text(_label("survey_summary", language), language, hindi_available),
        heading_style,
    ))
    summary_pairs = [
        ("metric", "value", True),  # header row
        ("total_frames", str(scan["total_frames"]), False),
        ("total_assets", str(scan["total_detections"]), False),
        ("signs_detected", str(scan["signs_detected"]), False),
        ("markings_detected", str(scan["markings_detected"]), False),
        ("studs_detected", str(scan["studs_detected"]), False),
        ("compliance_pass", str(scan["compliance_pass"]), False),
        ("compliance_fail", str(scan["compliance_fail"]), False),
        ("avg_score", f"{scan['avg_retro_score']:.1f} / 100", False),
    ]

    summary_data = []
    for key, val, _is_header in summary_pairs:
        if _is_header:
            label_text = _label("metric", language)
            value_text = _label("value", language)
        else:
            label_text = _label(key, language)
            value_text = val
        summary_data.append([
            Paragraph(_para_text(label_text, language, hindi_available), body_style),
            Paragraph(_para_text(value_text, language, hindi_available), body_style),
        ])

    summary_table = Table(summary_data, colWidths=[8 * cm, 8 * cm])
    summary_table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1a1a2e")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTNAME", (0, 0), (-1, 0), base_font_bold),
        ("FONTNAME", (0, 1), (-1, -1), base_font),
        ("FONTSIZE", (0, 0), (-1, -1), 10),
        ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
        ("ALIGN", (1, 0), (1, -1), "CENTER"),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#f5f5f5")]),
        ("LEFTPADDING", (0, 0), (-1, -1), 8),
        ("TOPPADDING", (0, 0), (-1, -1), 5),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 5),
    ]))
    elements.append(summary_table)
    elements.append(Spacer(1, 8 * mm))

    # Detailed findings
    elements.append(Paragraph(
        _para_text(_label("detailed_findings", language), language, hindi_available),
        heading_style,
    ))

    if detections:
        det_header = [
            "#",
            _label("asset_type", language),
            _label("retro_class", language),
            _label("score", language),
            _label("grade", language),
            _label("irc_status", language),
            _label("priority", language),
        ]
        det_rows = [[
            Paragraph(_para_text(h, language, hindi_available), body_style) for h in det_header
        ]]

        pass_label = "PASS" if language == "en" else (
            "उत्तीर्ण" if language == "hi" else "PASS / उत्तीर्ण"
        )
        fail_label = "FAIL" if language == "en" else (
            "विफल" if language == "hi" else "FAIL / विफल"
        )

        for i, det in enumerate(detections[:50], 1):  # Limit to 50 for PDF
            compliant = det["compliance"].get("compliant", False)
            row = [
                str(i),
                det["class_name"].replace("_", " ").title(),
                det["retro_class"],
                f"{det['retro_score']:.0f}",
                det["condition_grade"],
                pass_label if compliant else fail_label,
                str(det["priority"]),
            ]
            det_rows.append([
                Paragraph(_para_text(c, language, hindi_available), body_style) for c in row
            ])

        det_table = Table(det_rows, colWidths=[1 * cm, 4 * cm, 2.5 * cm, 1.5 * cm, 1.5 * cm, 2.5 * cm, 2 * cm])
        style_commands = [
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1a1a2e")),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
            ("FONTNAME", (0, 0), (-1, 0), base_font_bold),
            ("FONTNAME", (0, 1), (-1, -1), base_font),
            ("FONTSIZE", (0, 0), (-1, -1), 8),
            ("GRID", (0, 0), (-1, -1), 0.3, colors.grey),
            ("ALIGN", (0, 0), (-1, -1), "CENTER"),
            ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#f5f5f5")]),
            ("TOPPADDING", (0, 0), (-1, -1), 3),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 3),
        ]

        # Color-code retro class and IRC status cells (we color by raw class string)
        for row_idx in range(1, len(det_rows)):
            det = detections[row_idx - 1]
            retro_class = det["retro_class"]
            if retro_class == "HIGH":
                style_commands.append(("TEXTCOLOR", (2, row_idx), (2, row_idx), colors.HexColor("#2e7d32")))
            elif retro_class == "FAILED":
                style_commands.append(("TEXTCOLOR", (2, row_idx), (2, row_idx), colors.red))
            elif retro_class == "LOW":
                style_commands.append(("TEXTCOLOR", (2, row_idx), (2, row_idx), colors.HexColor("#e65100")))

            if not det["compliance"].get("compliant", False):
                style_commands.append(("BACKGROUND", (5, row_idx), (5, row_idx), colors.HexColor("#ffebee")))
                style_commands.append(("TEXTCOLOR", (5, row_idx), (5, row_idx), colors.red))

        det_table.setStyle(TableStyle(style_commands))
        elements.append(det_table)
    else:
        elements.append(Paragraph(
            _para_text(_label("no_assets", language), language, hindi_available),
            body_style,
        ))

    elements.append(Spacer(1, 10 * mm))

    # Recommendations
    elements.append(Paragraph(
        _para_text(_label("recommendations", language), language, hindi_available),
        heading_style,
    ))
    critical = [d for d in detections if d["priority"] <= 2]
    if critical:
        for det in critical[:10]:
            issues = det.get("issues", [])
            issue_text = "; ".join(issues) if isinstance(issues, list) else str(issues)
            asset_label = det['class_name'].replace('_', ' ').title()
            line = (
                f"<b>{html_escape(asset_label)}</b> "
                f"({_label('frame', language)} {det['frame_index']}) — "
                f"{_label('score', language)}: {det['retro_score']:.0f}, "
                f"{_label('priority', language)}: {html_escape(str(det['priority_label']))}. "
                f"{_label('issues', language)}: {html_escape(issue_text)}"
            )
            if language in {"hi", "bilingual"} and hindi_available:
                line = f'<font name="{HINDI_FONT_NAME}">{line}</font>'
            elements.append(Paragraph(line, body_style))
            elements.append(Spacer(1, 3 * mm))
    else:
        elements.append(Paragraph(
            _para_text(_label("no_critical", language), language, hindi_available),
            body_style,
        ))

    # Footer
    elements.append(Spacer(1, 15 * mm))
    footer_text = _label("footer", language)
    footer_html = f"<i>{html_escape(footer_text)}</i>"
    if language in {"hi", "bilingual"} and hindi_available:
        footer_html = f'<font name="{HINDI_FONT_NAME}"><i>{html_escape(footer_text)}</i></font>'
    elements.append(Paragraph(
        footer_html,
        ParagraphStyle(
            "Footer", parent=body_style, fontSize=8, textColor=colors.grey,
            fontName=base_font,
        ),
    ))

    doc.build(elements)
    return buffer.getvalue()


def generate_excel_report(scan_id: str) -> bytes:
    """Generate an Excel data export."""
    scan = get_scan(scan_id)
    detections = get_scan_detections(scan_id)

    if not scan:
        raise ValueError(f"Scan {scan_id} not found")

    wb = openpyxl.Workbook()

    # Summary sheet
    ws_summary = wb.active
    ws_summary.title = "Summary"

    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="1A1A2E", end_color="1A1A2E", fill_type="solid")
    pass_fill = PatternFill(start_color="E8F5E9", end_color="E8F5E9", fill_type="solid")
    fail_fill = PatternFill(start_color="FFEBEE", end_color="FFEBEE", fill_type="solid")
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )

    summary_rows = [
        ("Metric", "Value"),
        ("Scan ID", scan_id),
        ("File", scan["filename"]),
        ("Date", scan["created_at"]),
        ("Total Frames", scan["total_frames"]),
        ("Total Detections", scan["total_detections"]),
        ("Signs", scan["signs_detected"]),
        ("Markings", scan["markings_detected"]),
        ("Studs", scan["studs_detected"]),
        ("Compliance Pass", scan["compliance_pass"]),
        ("Compliance Fail", scan["compliance_fail"]),
        ("Avg Retro Score", scan["avg_retro_score"]),
    ]

    for row_idx, (metric, value) in enumerate(summary_rows, 1):
        ws_summary.cell(row=row_idx, column=1, value=metric).border = thin_border
        ws_summary.cell(row=row_idx, column=2, value=value).border = thin_border
        if row_idx == 1:
            ws_summary.cell(row=row_idx, column=1).font = header_font
            ws_summary.cell(row=row_idx, column=1).fill = header_fill
            ws_summary.cell(row=row_idx, column=2).font = header_font
            ws_summary.cell(row=row_idx, column=2).fill = header_fill

    ws_summary.column_dimensions["A"].width = 20
    ws_summary.column_dimensions["B"].width = 40

    # Detections sheet
    ws_det = wb.create_sheet("Detections")
    headers = [
        "ID", "Frame", "Timestamp (s)", "Asset Type", "Category",
        "Confidence", "Retro Class", "Retro Score", "Grade",
        "Color Fading", "Surface Damage", "Dirt Level", "Legibility",
        "IRC Standard", "Compliant", "Threshold", "Estimated Value",
        "Priority", "Priority Label", "Cost Estimate (INR)",
        "Issues", "Age Estimate",
    ]

    for col, header in enumerate(headers, 1):
        cell = ws_det.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")
        cell.border = thin_border

    for row_idx, det in enumerate(detections, 2):
        compliance = det.get("compliance", {})
        issues = det.get("issues", [])
        values = [
            det["id"], det["frame_index"], round(det["timestamp_sec"], 2),
            det["class_name"].replace("_", " ").title(), det["category"],
            det["confidence"], det["retro_class"], det["retro_score"],
            det["condition_grade"], det["color_fading"], det["surface_damage"],
            det["dirt_level"], det["legibility"],
            compliance.get("standard", ""), "PASS" if compliance.get("compliant") else "FAIL",
            compliance.get("threshold", ""), compliance.get("estimated_ra", compliance.get("estimated_rl", "")),
            det["priority"], det["priority_label"], det["cost_estimate"],
            "; ".join(issues) if isinstance(issues, list) else str(issues),
            det["age_estimate"],
        ]

        for col, value in enumerate(values, 1):
            cell = ws_det.cell(row=row_idx, column=col, value=value)
            cell.border = thin_border
            cell.alignment = Alignment(horizontal="center")

        # Color-code compliance column
        compliance_cell = ws_det.cell(row=row_idx, column=15)
        if compliance.get("compliant"):
            compliance_cell.fill = pass_fill
        else:
            compliance_cell.fill = fail_fill

    # Auto-width columns
    for col in range(1, len(headers) + 1):
        ws_det.column_dimensions[openpyxl.utils.get_column_letter(col)].width = 15

    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue()
